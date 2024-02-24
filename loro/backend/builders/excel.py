#!/usr/bin/python
# File: EWA.py
# Author: Tomás Vírseda
# Description: This script cleans up EWAs_Performance_overview excel

import sys
import pprint

import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )

from loro.backend.core.env import ENV
from loro.backend.services.nlp.spacy import get_glossary_term_explained

def as_text(value):
    if value is None:
        return ""
    return str(value)

def create_excel(stats, posset):
    wb = Workbook()
    ws = wb.active
    wb.remove_sheet(ws)
    header = []
    for key in stats:
        if key in posset:
            postag = get_glossary_term_explained(key).title()
            ws = wb.create_sheet(postag)
            filters = ws.auto_filter
            header = ['Word', 'Count']
            data = []
            data.append(header)
            for word, count in sorted(stats[key].items(), key=lambda p:p[1], reverse=True):
                data.append([word, count])
            for row in data:
                ws.append(row)
            a1 = ws['A1']
            b1 = ws['B1']
            a1.font = Font(name='Arial', size=12, bold=True)
            b1.font = Font(name='Arial', size=12, bold=True)

            for column_cells in ws.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length

    try:
        wb.save("/tmp/topdict.xlsx")
    except IndexError:
        print("No excel workbook created")

